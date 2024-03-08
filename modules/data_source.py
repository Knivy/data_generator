"""Модуль, получающий данные и отдающий построчно списками строк."""

import xlrd  # type: ignore[import-untyped]
import openpyxl  # type: ignore[import-untyped]

from .generator import MaleFemaleFakerGenerator, MaleFemaleMimesisGenerator
from .generator import BaseGenerator


class BaseDataSource:
    """Базовый класс источников данных."""

    default_num_lines: int = 500_000  # Сколько нужно строк.

    def __init__(self, num_lines: int = default_num_lines):
        """Получает число строк."""
        self.num_lines = num_lines

    def get_data(self):
        """Получает данные и отдает построчно."""
        raise NotImplementedError('Этот метод должен быть переопределён.')


class InputDataSource(BaseDataSource):
    """Данные от ввода пользователя."""

    def get_data(self):
        """Получает данные от пользователя и отдает построчно."""

        print('Введите данные построчно через пробел.',
              'Enter - окончание ввода.')
        for _ in range(self.num_lines):
            while True:
                line = input('Введите строку: ')
                if line:
                    break
            yield line.split()


class GeneratorDataSource(BaseDataSource):
    """Данные от генератора синтетических данных."""

    default_generator_type = 'Mimesis'

    def __init__(self,  num_lines: int = BaseDataSource.default_num_lines,
                 language: str = BaseGenerator.default_language,
                 format: str = BaseGenerator.default_format,
                 header: bool = BaseGenerator.header,
                 generator_type: str = default_generator_type):

        super().__init__(num_lines)
        self.header = header
        if generator_type.lower() != self.default_generator_type:
            self.generator: BaseGenerator = MaleFemaleFakerGenerator(language,
                                                                     format)
        else:
            self.generator = MaleFemaleMimesisGenerator(language, format)

    def get_data(self):
        """Получает данные от генератора."""
        for line in self.generator.generate(self.num_lines, self.header):
            yield line


class FileDataSource(BaseDataSource):
    """Данные из файла."""

    file_formats: tuple

    def __init__(self, file_path: str,
                 num_lines: int = BaseDataSource.default_num_lines):
        """Получает настройки."""
        super().__init__(num_lines)
        self.file_path = file_path


class CsvDataSource(FileDataSource):
    """Данные из файла csv/txt."""

    file_formats: tuple = ('csv', 'txt')
    default_separator: str = ','

    def __init__(self, file_path: str,
                 num_lines: int = BaseDataSource.default_num_lines,
                 separator: str = default_separator):
        """Получает настройки."""
        super().__init__(file_path, num_lines)
        self.separator = separator

    def get_data(self):
        """Получает данные из файла txt или csv."""

        with open(self.file_path, 'r', encoding='utf-8') as f:
            line_count = 0
            for line in f.readlines():
                line_count += 1
                if line_count > self.num_lines:
                    raise StopIteration
                line = line.strip()
                if line:
                    yield line.split(self.separator)


class XlsxDataSource(FileDataSource):
    """Данные из первого листа xlsx-файла."""

    file_formats = ('xlsx',)

    def get_data(self):
        """Получает данные из файла xlsx."""
        workbook = openpyxl.load_workbook(self.file_path)
        worksheet = workbook.active
        length = worksheet.max_row
        num_lines = min(length, self.num_lines)
        for row in range(num_lines):
            line = []
            for col in worksheet.iter_cols(1, worksheet.max_column):
                line.append(col[row].value)
            yield line


class XlsDataSource(FileDataSource):
    """
    Данные из xls-файла.
    """

    file_formats = ('xls',)
    default_sheet_number = 0

    def __init__(self, file_path: str,
                 num_lines: int = BaseDataSource.default_num_lines,
                 sheet_number: int = default_sheet_number):
        """Получает настройки."""
        super().__init__(file_path, num_lines)
        self.sheet_number = sheet_number

    def get_data(self):
        """Получает данные из файла xls."""
        workbook = xlrd.open_workbook(self.file_path)
        worksheet = workbook.sheet_by_index(self.sheet_number)
        length = worksheet.nrows
        num_lines = min(length, self.num_lines)
        for row in range(num_lines):
            line = []
            for col in range(worksheet.ncols):
                line.append(worksheet.cell_value(row, col))
            yield line
