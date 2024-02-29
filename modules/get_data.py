"""Модуль, получающий данные и отдающий построчно списками строк."""

from .generator import MaleFemaleGenerator
import xlrd
import openpyxl


class WrongExtension(Exception):
    """Класс пользовательского исключения при неверном расширении."""

    def __str__(self):
        return 'Неверное расширение файла.'


class DataSource:
    """Базовый класс источников данных."""

    def get_data(self):
        """Получает данные и отдает построчно."""
        raise NotImplementedError

    def get_num_lines(self):
        """Запрашивает число строк."""
        while True:
            try:
                num_lines = int(input('Введите число строк: '))
            except ValueError:
                print('Неверный ввод.')
            if num_lines >= 0:
                break
            else:
                print('Введите неотрицательное число.')
        return num_lines


class InputDataSource(DataSource):
    """Данные от ввода пользователя."""

    def get_data(self):
        """Получает данные от пользователя."""
        num_lines = self.get_num_lines()
        print('Введите данные построчно через пробел.',
              'Enter - окончание ввода.')
        for _ in range(num_lines):
            line = input('Введите строку: ')
            if not line:
                continue
            yield line.split()


class GeneratorDataSource(DataSource):
    """Данные от генератора."""

    def get_data(self):
        """Получает данные от генератора."""
        num_lines = self.get_num_lines()
        lang = input('Введите язык (по умолчанию ru_RU): ') or 'ru_RU'
        format = input('Введите формат male/female/general: ') or 'general'
        fk = MaleFemaleGenerator(lang, format)
        for line in fk.generate(num_lines):
            yield line


class FileDataSource(DataSource):
    """Данные из файла."""

    def get_data(self):
        """Получает данные из файла txt или csv."""
        file_path = input('Введите путь к файлу через /: ')
        sep = input('Введите символ разделителя: ')
        with open(file_path, 'r', encoding='utf-8') as f:
            length = len(f.readlines())
            num_lines = min(length, self.get_num_lines())
            f.seek(0)
            for line in f.readlines()[:num_lines]:
                line = line.strip()
                if line:
                    yield line.split(sep)


class XlsxDataSource(FileDataSource):
    """Данные из xlsx-файла."""

    def get_data(self):
        """Получает данные из файла xlsx."""
        file_path = input('Введите путь к файлу через /: ')
        if file_path.split('.')[-1] not in ['xlsx']:
            raise WrongExtension
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        length = worksheet.max_row
        num_lines = min(length, self.get_num_lines())
        for row in range(num_lines):
            line = []
            for col in worksheet.iter_cols(1, worksheet.max_column):
                line.append(col[row].value)
            yield line


class XlsDataSource(FileDataSource):
    """
    Данные из xls-файла.

    Не тестировалось, так как под Линуксом нет офиса.
    """

    def get_data(self):
        """Получает данные из файла xls."""
        file_path = input('Введите путь к файлу через /: ')
        if file_path.split('.')[-1] not in ['xls']:
            raise WrongExtension
        workbook = xlrd.open_workbook(file_path)
        sheet = int(input('Введите номер листа: '))
        worksheet = workbook.sheet_by_index(sheet)
        length = worksheet.nrows
        num_lines = min(length, self.get_num_lines())
        for row in range(num_lines):
            line = []
            for col in range(worksheet.ncols):
                line.append(worksheet.cell_value(row, col))
            yield line
